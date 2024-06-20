from datetime import datetime
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect
from django.http import HttpResponse 

from proyecto.models import Catorcenas, Variables_imss_patronal, SalarioDatos, TablaVacaciones, DatosISR, TablaSubsidio
from revisar.models import AutorizarPrenomina
from esquema.models import BonoSolicitado
from prenomina.models import PrenominaIncidencias, TipoAguinaldo, Aguinaldo

from datetime import timedelta, datetime
from decimal import Decimal
import calendar
from dateutil.relativedelta import relativedelta

#CUOTAS IMSS
def calcular_cuotas_imss(request,sdi_imss):
    variables_patronal = Variables_imss_patronal.objects.get()
    salario_datos = SalarioDatos.objects.get()
    
    #multiplica el sdi * el % de cuatoas / el numero de dias de la catorcena
    invalidez_vida = sdi_imss * Decimal(variables_patronal.iv_obrero / 100) * 14
    cesantia_vejez = sdi_imss * Decimal(variables_patronal.cav_patron/100) * 14
    
    #obtener el salario cotizacion mensual
    salario_cot_men = sdi_imss * Decimal(30.4)
    gastos_medicos = sdi_imss * Decimal(variables_patronal.gmp_obrero/100) * 14
    en_dinero = sdi_imss * Decimal(variables_patronal.pd_obrero/100) * 14 
    
    #calcular cuota fija por cada trabajador hasta por 3 UMAs
    cuota_fija_umas = salario_datos.UMA * Decimal(30.4) * 3
    diferencia_sbc_umas = salario_cot_men - cuota_fija_umas
    cuota_fija = (diferencia_sbc_umas * Decimal(variables_patronal.cf_obrero / 100) / Decimal(30.4)) * 14
    enfermedades_maternidad = gastos_medicos + en_dinero + cuota_fija
    
    #La suma del calculo de cada resultado    
    calculo_imss = invalidez_vida + cesantia_vejez + enfermedades_maternidad
    print("Este es el calculo IMSS: ", calculo_imss)
    return calculo_imss

#CALULAR ISR
def calcular_isr(request,salario,prima_dominical_isr,calulo_aguinaldo_isr,calculo_aguinaldo_eventual_isr):   
    salario_datos = SalarioDatos.objects.get()
    limite_inferior = 0
    porcentaje = 0
    cuota_fija = 0
    #Salario minino queda exento
    if salario > salario_datos.Salario_minimo:
        #PRIMA DOMINICAL
        if prima_dominical_isr < salario_datos.UMA:
            #exento
            prima_dominical_isr = 0
            print("prima dominical exenta: ", prima_dominical_isr)
        else:
            #gravable
            prima_dominical_isr = prima_dominical_isr - Decimal(salario_datos.UMA)
            print("prima dominical gravable: ", prima_dominical_isr)

        #AGUINALDO
        if calulo_aguinaldo_isr < (salario_datos.UMA * 30):
            #exento
            calulo_aguinaldo_isr = 0
        else:
            #gravado
            calulo_aguinaldo_isr - Decimal(salario_datos.UMA * 30)
            
            
        #AGUINALDO EVENTUAL
        if calculo_aguinaldo_eventual_isr <  (salario_datos.UMA * 30):
            #exento
            calculo_aguinaldo_eventual_isr = 0
            print("aguinaldo envetual exento", calculo_aguinaldo_eventual_isr)
        else:
            #gravado
            calculo_aguinaldo_eventual_isr - Decimal(salario_datos.UMA * 30)
            print("aguinaldo envetual gravado", calculo_aguinaldo_eventual_isr)
        
    else:
        #exento
        prima_dominical_isr = 0
        calulo_aguinaldo_isr = 0
        calculo_aguinaldo_eventual_isr = 0
        
    #multiplicar el salario por 30.4
    salario_catorcenal = salario * Decimal(salario_datos.dias_mes) #30.4
    #se suman la prima dominical, vacacional, los aguinaldos para despues aplicar el calculo del isr
    #salario_catorcenal = salario_catorcenal + prima_dominical_isr + calulo_aguinaldo_isr + calculo_aguinaldo_eventual_isr

     
    #llamar la tabla de IRS
    tabla_irs = DatosISR.objects.all()
    
    #obtener el valor aproximado hacia abajo para obtener las variables
    for datos_irs in tabla_irs:
        if salario_catorcenal >= datos_irs.liminf:
            limite_inferior = datos_irs.liminf
            porcentaje = datos_irs.excedente
            cuota_fija = datos_irs.cuota
            
    #realizar el calculo
    isr_mensual = ((salario_catorcenal - limite_inferior) * porcentaje) + cuota_fija
    
    isr_catorcenal = (isr_mensual / salario_datos.dias_mes) * 14
    
    print("calculo ISR: ", isr_mensual)
    
    return isr_catorcenal

#OBTENER EL TOTAL DE BONOS POR EMPLEADO Y CATORCENA
def obtener_total_bonos(request,prenomina, catorcena):
    #Fecha para obtener los bonos agregando la hora y la fecha de acuerdo a la catorcena
    fecha_inicial = datetime.combine(catorcena.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
    fecha_final = datetime.combine(catorcena.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)
    
    total_bonos = BonoSolicitado.objects.filter(
        trabajador_id=prenomina.empleado.status.perfil.id,
        solicitud__fecha_autorizacion__isnull=False,
        solicitud__fecha_autorizacion__range=(fecha_inicial, fecha_final)
    ).aggregate(total=Sum('cantidad'))['total'] or 0

    return total_bonos

#CALCULAR INFONAVIT
def calcular_infonavit(request, infonavit):
    if infonavit == 0:
        prestamo_infonavit = Decimal(0.00)
    else:
        prestamo_infonavit = Decimal((infonavit / Decimal(30.4) ) * 14 )
    
    return prestamo_infonavit 

#CALCULAR FONACOT
def calcular_fonacot(request,fonacot):
    if fonacot == 0:
        prestamo_fonacot = Decimal(0.00)
    else:
        #Se haya la catorcena actual, y cuenta cuantas catorcenas le corresponden al mes actual
        primer_dia_mes = datetime(datetime.now().year, datetime.now().month, 1).date()
        ultimo_dia_mes = datetime(datetime.now().year, datetime.now().month,
                                calendar.monthrange(datetime.now().year, datetime.now().month)[1]).date()
        numero_catorcenas =  Catorcenas.objects.filter(fecha_final__range=(primer_dia_mes,ultimo_dia_mes)).count()
        prestamo_fonacot = prestamo_fonacot / numero_catorcenas
        
    return prestamo_fonacot

#PRIMA DOMINICAL
def calcular_prima_dominical(request,dia_extra,salario):
    dato = SalarioDatos.objects.get()
    prima_dominical = salario * Decimal(dato.prima_vacacional)
    prima_dominical = prima_dominical * dia_extra
    print("salario", salario)
    print("prima dominical", Decimal(prima_dominical))
    return Decimal(prima_dominical)

#PRIMA VACACIONAL
def calcular_prima_vacacional(request,salario,prenomina):
    tabla_vacaciones = TablaVacaciones.objects.all()
    dato = SalarioDatos.objects.get()
    
    fecha_actual = datetime.today()
       
    calcular_prima = True
    if prenomina.empleado.status.tipo_de_contrato_id == 4: #HONORARIOS
        calcular_prima = False
    elif prenomina.empleado.status.tipo_de_contrato_id == 2: #EVENTUAL
        fecha = fecha_actual - timedelta(days=365) # El calculo es 12 dias de vacaciones, siempre para contrato eventual
    elif prenomina.empleado.status.tipo_de_contrato_id == 7: #NR
        calcular_prima = False
    elif prenomina.empleado.status.fecha_planta is None and prenomina.empleado.status.fecha_planta_anterior is None:
        calcular_prima = False
    elif prenomina.empleado.status.fecha_planta is not None and prenomina.empleado.status.fecha_planta_anterior is not None:
        fecha = prenomina.empleado.status.fecha_planta_anterior
    elif prenomina.empleado.status.fecha_planta:
        fecha = prenomina.empleado.status.fecha_planta
    elif prenomina.empleado.status.fecha_planta_anterior:
        fecha = prenomina.empleado.status.fecha_planta_anterior

    prima_vacacional = 0
    if calcular_prima == True:#calcula la prima
        calcular_antiguedad = relativedelta(fecha_actual, fecha)
        antiguedad = calcular_antiguedad.years
        print("esta es la antiguedad ", antiguedad)

        if antiguedad > 0:
            for tabla in tabla_vacaciones:
                if antiguedad >= tabla.years:
                    dias_vacaciones = tabla.days #Se asignan los días para el calculo de la prima vacacional

            vac_reforma_actual = Decimal(dias_vacaciones) * Decimal(salario)
            print("dias vacaciones", dias_vacaciones, "salario", salario)
            print("vacaciones", vac_reforma_actual)
            
            prima_vacacional = vac_reforma_actual*Decimal(dato.prima_vacacional)
            print("esta es la prima ", prima_vacacional)
            return prima_vacacional

        else:#No calcula la prima - No tiene el año de antiguedad o más
            print("esta es la prima ", prima_vacacional)
            return prima_vacacional

    else:#No calcula la prima
        print("esta es la prima ", prima_vacacional)
        return prima_vacacional  

#Calcular aguinaldo enventual
def calcular_aguinaldo_eventual(request,salario,prenomina):
    """
    se realiza el calculo del registro cuando cumpla el tiempo y se guarda en la base de datos, para guardar
    se debe estar en complete = True (pagado), mes (el ultimo mes del contrato) y se considera que se va 
    a pagar en la siguiente catorcena
    """
    #tipo contrato
    tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id    
    #el mes corresponde cuando cumple 1°, 2° o 3° mes
    mes = 0
    aguinaldo = Decimal(0.00)
    if tipo_contrato == 2:#eventual
        #fecha ingreso
        fecha_ingreso =  prenomina.empleado.status.fecha_ingreso
        #calculo relacion dias laborados
        calculo_fecha = relativedelta(datetime.date.today(),fecha_ingreso)
        
        #se lleva un registro de los aguinaldos que se van registrando 
        aguinaldo_registrado = Aguinaldo_Contrato.objects.filter(empleado_id = prenomina.empleado.id).last()
        
        if aguinaldo_registrado is None:
            #se realiza el calculo por los meses y por los dias laborados correspondientes a cada condicion
            if calculo_fecha.months >= 1 and calculo_fecha.days >= 0:
                #primer mes laborado le corresponden 30 dias
                dias_aguinaldo = Decimal(30 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 1
        else:
            if aguinaldo_registrado.mes == 1 and aguinaldo_registrado.complete == True and calculo_fecha.months >= 3 and calculo_fecha.days >= 0:
                #tercer mes laborado le corresponden 60 dias
                dias_aguinaldo = Decimal(60 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 3
            elif aguinaldo_registrado.mes == 3 and aguinaldo_registrado.complete == True and calculo_fecha.months >= 6 and calculo_fecha.days >= 0:
                #sexto mes laborado le corresponden 90 dias
                dias_aguinaldo = Decimal(90 * 15) / 365
                aguinaldo = dias_aguinaldo * salario
                mes = 6
            print("este es el aguinaldo eventual: ",aguinaldo)
            #se guarda en caso que exista un valor en el aguinaldo
        
        if aguinaldo != 0:
            #saber la catorcena actual
            ahora = datetime.date.today()
            catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
            #Guardar el aguinaldo
            aguinaldo_contrato = Aguinaldo_Contrato(
                empleado = prenomina.empleado,
                aguinaldo = aguinaldo,
                fecha = date.today(),
                catorcena = catorcena_actual.id + 1, #el aguinaldo se paga en la cat siguiente
                complete=False,
                mes = mes
            )
            aguinaldo_contrato.save()

#CALCULAR AGUINALDO ANUAL O PROPORCIONAL depende fecha
def calcular_aguinaldo(request,salario,prenomina):
    aguinaldo = Decimal(0.00)
    #obtener la catorcena actual - se utiliza para comparar la catorcena cuando sea diciembre
    ahora = datetime.today()
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()
    print("esta es la catorcena para comparar: ",catorcena_actual.id)
    #obtener la primer cat de diciembre fecha y dia 
    catorcena_decembrina = Catorcenas.objects.filter(fecha_inicial__month=12, fecha_final__month=12,fecha_inicial__year=ahora.year,fecha_final__year=ahora.year).first()
    #print("esta es la catorcena de diciembre", catorcena_decembrina.id)
    
    #verifica si es la catorcena de diciembre para pagar el aguinaldo
    #if catorcena_actual.id == catorcena_decembrina.id:
    if catorcena_actual.id == 1000:
        
        tipo_contrato = prenomina.empleado.status.tipo_de_contrato_id
        
        if tipo_contrato in (1,3,5,6): # planta, especial, planta 1, planta 2
            fecha_planta = prenomina.empleado.status.fecha_planta
            fecha_planta_anterior = prenomina.empleado.status.fecha_planta_anterior
                    
            #Se obtiene la fecha de planta o planta anterior
            if fecha_planta is None and fecha_planta_anterior is None:
                fecha = None
            elif fecha_planta_anterior is not None and fecha_planta is not None:
                fecha = fecha_planta_anterior
            elif fecha_planta:
                fecha = fecha_planta
            elif fecha_planta_anterior:
                fecha = fecha_planta_anterior
                        
            #de acuerdo a la fecha se realiza el calculo para el aguinaldo
            if fecha is not None:
                fecha = datetime.date(2022, 12, 31)
                
                
                antiguedad = relativedelta(datetime.date.today(),fecha)
            
                if antiguedad.years >= 1:#Aguinaldo completo >= al 1 primer año de antiguedad
                    #aqui es con respecto al año 1/ENE - 31/DIC
                    año_actual = datetime.date.today().year
                    inicio_año = datetime.date(año_actual, 1, 1)
                    fin_año = datetime.date(año_actual, 12, 31)
                    
                    #llama funcion para el conteo de las incidencias necesarias para el aguinaldo
                    total_incidencias = conteo_incidencias_aguinaldo(request,prenomina,inicio_año,fin_año)
                    #se restan las incidencias con los dias laborados proporcionales
                    dias_laborados = 365 - total_incidencias
                    
                    dias_pago = Decimal((dias_laborados) * 15 / 365)
                    aguinaldo = Decimal(dias_pago * salario)
                    print("aguinaldo completo")
                    print(aguinaldo)
                    #exit()
                    return aguinaldo
                else: #aguinaldo proporcional 
                    #No cumple el año y se obtiene el proporcional
                    fecha_final = datetime.date(datetime.date.today().year, 12, 31) #obtener fin de año
                    diferencia = fecha_final - fecha 
                    dias_aguinaldo = diferencia.days #total de dias laborados
                    #fecha es la fecha de planta o anterior
                    dias_incidencias = conteo_incidencias_aguinaldo(request,prenomina,fecha,fecha_final)
                
                    #se restan las incidencias con los dias laborados proporcionales
                    dias_laborados = dias_aguinaldo - dias_incidencias
                            
                    dias_pago = Decimal((dias_laborados * 15)/365)
                    print("aguinaldo proporcional ")
                    aguinaldo = Decimal( dias_pago * salario)
                    print(aguinaldo)
                    #exit()
                    return aguinaldo
            else:
                return aguinaldo
        else:
            return aguinaldo
    else:
        print("no corresponde pago del aguinaldo")
        return aguinaldo

def calcular_subsidio(request, salario_catorcenal,):
    tabla_subsidio = TablaSubsidio.objects.all()
    for dato in tabla_subsidio:
        if salario_catorcenal >= dato.liminf:
            subsidio=dato.cuota

    return (subsidio)

def calcular_incidencias(request, prenomina, catorcena_actual):
    # Contadores para cada tipo de incidencia
    retardos = 0
    descansos = 0
    faltas = 0
    comisiones = 0
    domingos = 0
    dia_extra = 0 #dia de descanso laborado
    festivos = 0
    economicos = 0
    castigos = 0
    permisos_sin_goce = 0
    permisos_con_goce = 0
    vacaciones = 0
    incapacidad_enfermedad_general = 0  # Excluido
    incapacidad_riesgo_laboral = 0
    incapacidad_maternidad = 0

    # Obtener todas las incidencias en una sola consulta
    incidencias = PrenominaIncidencias.objects.filter(
        prenomina__empleado_id=prenomina.empleado_id,
        fecha__range=(catorcena_actual.fecha_inicial, catorcena_actual.fecha_final)
    )

    # Contar las incidencias por tipo
    for incidencia in incidencias:
        incidencia = incidencia.incidencia.id
        if incidencia == 1:
            retardos += 1
        elif incidencia == 2:
            descansos += 1
        elif incidencia == 3:
            faltas += 1
        elif incidencia == 4:
            comisiones += 1
        elif incidencia == 5:
            domingos += 1
        elif incidencia == 6: #dia de descanso laborado
            dia_extra += 1
        elif incidencia == 7:
            castigos += 1
        elif incidencia == 8:
            permisos_sin_goce += 1
        elif incidencia == 9:
            permisos_con_goce += 1
        elif incidencia == 10:
            incapacidad_enfermedad_general += 1
        elif incidencia == 11:
            incapacidad_riesgo_laboral += 1
        elif incidencia == 12:
            incapacidad_maternidad += 1
        elif incidencia == 13:
            festivos += 1
        elif incidencia == 14:
            economicos += 1
        elif incidencia == 15:
            vacaciones += 1

    return (retardos, descansos, faltas, comisiones, domingos, dia_extra, castigos, 
            permisos_sin_goce, permisos_con_goce, incapacidad_riesgo_laboral, 
            incapacidad_maternidad, festivos, economicos, vacaciones, incapacidad_enfermedad_general)

#GENERAR REPORTE PRENOMINA ACTUAL
def excel_estado_prenomina(request,prenominas, user_filter):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_prenominas_' + str(datetime.now())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1

    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
    date_style.font = Font(name ='Calibri', size = 11)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 11)
    bold_money_style = NamedStyle(name='bold_money_style', number_format='$#,##0.00', font=Font(bold=True))
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
        
    columns = ['Empleado','#Trabajador','Distrito','#Catorcena','Fecha','Estado general','RH','CT','Gerencia','Autorizada','Retardos','Castigos','Permiso con goce de sueldo',
               'Permiso sin goce de sueldo','Descansos','Incapacidades','Faltas','Comisión','Domingo','Dia de descanso laborado','Festivos','Economicos','Vacaciones','Salario Cartocenal',
               'Previsión social', 'Total bonos','Total percepciones','Prestamo infonavit','IMSS','Fonacot','ISR Retenido','Total deducciones','Neto a pagar en nomina']

    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num == 1:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 50
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15


    columna_max = len(columns)+2
    
    ahora = datetime.now()
    #ahora = datetime.now() + timedelta(days=15)
    
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=ahora, fecha_final__gte=ahora).first()

    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. JH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    #(ws.cell(column = columna_max, row = 3, value='Algún dato')).style = messages_style
    #(ws.cell(column = columna_max +1, row=3, value = 'alguna sumatoria')).style = money_resumen_style
    (ws.cell(column = columna_max, row = 4, value=f'Catorcena: {catorcena_actual.catorcena}: {catorcena_actual.fecha_inicial.strftime("%d/%m/%Y")} - {catorcena_actual.fecha_final.strftime("%d/%m/%Y")}')).style = dato_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 50
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 50

    rows = []

    sub_salario_catorcenal_costo = Decimal(0.00) #Valor de referencia del costo
    sub_salario_catorcenal = Decimal(0.00)
    sub_apoyo_pasajes = Decimal(0.00)
    sub_total_bonos = Decimal(0.00)
    sub_total_percepciones = Decimal(0.00)
    sub_prestamo_infonavit = Decimal(0.00)
    sub_calculo_isr = Decimal(0.00)
    sub_calculo_imss = Decimal(0.00)
    sub_prestamo_fonacot = Decimal(0.00)
    sub_total_deducciones = Decimal(0.00)
    sub_pagar_nomina = Decimal(0.00)
        
    for prenomina in prenominas:        
        RH = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="RH").first()
        CT = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Control Tecnico").first()
        G = AutorizarPrenomina.objects.filter(prenomina=prenomina, tipo_perfil__nombre="Gerencia").first()

        if G is not None and G.estado.tipo == 'aprobado':
            estado = 'aprobado'
        elif G is not None and G.estado == 'rechazado':
            estado = 'rechazado'
        else:
            estado = 'pendiente'

        if RH is None:
            RH ="Ninguno"   
        else:
            RH = str(RH.perfil.nombres)+(" ")+str(RH.perfil.apellidos)
        if CT is None:
            CT ="Ninguno"
        else:
            CT = str(CT.perfil.nombres)+(" ")+str(CT.perfil.apellidos)
        if G is None:
            G ="Ninguno"
        else:
            G = str(G.perfil.nombres)+(" ")+str(G.perfil.apellidos)
        
        #datos para obtener los calculos de la prenomina dependiendo el empleado
        salario = Decimal(prenomina.empleado.status.costo.sueldo_diario)
        apoyo_pasajes = prenomina.empleado.status.costo.apoyo_de_pasajes
        infonavit = prenomina.empleado.status.costo.amortizacion_infonavit
        fonacot = prenomina.empleado.status.costo.fonacot 
        sdi_imss = prenomina.empleado.status.costo.sdi_imss
                
        #realiza el calculo de las cuotas imss
        calculo_imss = calcular_cuotas_imss(request,sdi_imss)
                
        #realiza el calculo de la prima vacacional
        calulo_prima_vacacional = calcular_prima_vacacional(request,salario,prenomina)
        
        #Obtener el total de los bonos
        total_bonos = obtener_total_bonos(request,prenomina,catorcena_actual)
                           
        #calculo del infonavit
        prestamo_infonavit = calcular_infonavit(request,infonavit)
                                        
        #calculo del fonacot
        prestamo_fonacot = calcular_fonacot(request,fonacot)
        
        print("EMPLEADO: ",prenomina.empleado)
        print("infonavit", prestamo_infonavit)
        print("fonacot", prestamo_fonacot)
        print("IMSS", calculo_imss)
        print("salario: ",salario)
        
        #Extrar el conteo de las incidencias de la funcion      
        retardos, descansos, faltas, comisiones, domingos, dia_extra, castigos, permisos_sin_goce, permisos_con_goce, incapacidad_riesgo_laboral, incapacidad_maternidad, festivos, economicos, vacaciones, incapacidad_enfermedad_general = calcular_incidencias(request, prenomina, catorcena_actual)
        #cont_riesgo,cont_enfermedad,pagar_dias_incapacidad,cont_maternidad,tipo = calcular_incapacidades(request,prenomina,catorcena_actual)
       
        #calculo de la prima se manda a llamar
        #if vacaciones > 0:
        #    cantidad_dias_vacacion = vacaciones
        #    calcular_prima_vacacional(request, cantidad_dias_vacacion)
            
        
        #numero de catorena
        catorcena_num = catorcena_actual.catorcena 
        
        incidencias = 0
        incidencias_retardos = 0
        
        if faltas > 0:
            incidencias = incidencias + faltas
            
        if retardos > 0:
            incidencias_retardos = retardos // 3 # Tres retardos se descuenta 1 dia
            
        if castigos > 0:
            incidencias = incidencias + castigos
            
        if permisos_sin_goce  > 0:
            incidencias = incidencias + permisos_sin_goce 
            
        prima_dominical = 0
        pago_doble = 0  
        if dia_extra > 0:
            pago_doble = Decimal(dia_extra * (salario * 2))
            prima_dominical = calcular_prima_dominical(request,dia_extra,salario)
            
        #calcular aguinaldo - siempre de ejecutara al momento de generar al reporte - verifica si es diciembre
        calulo_aguinaldo = calcular_aguinaldo(request,salario,prenomina)
         
        #calcular aguinaldo eventual - siempre se ejecutara al momento de generar el reporte
        #calcular_aguinaldo_eventual(request,salario,prenomina)
        
        #se realiza el calculo del aguinaldo por el tiempo del contrato, se calcula en una catorcena y se paga en la siguiente
        """
        aguinaldo_contrato = Aguinaldo_Contrato.objects.filter(empleado_id=prenomina.empleado.id).last()
        calculo_aguinaldo_eventual = 0
        if aguinaldo_contrato is not None and catorcena_actual.id == aguinaldo_contrato.catorcena:
            calculo_aguinaldo_eventual = aguinaldo_contrato.aguinaldo # se pasa el valor del aguinaldo del contrato para ser calculado en el ISR
            aguinaldo_contrato.complete = True #se actualiza el registro para definir que se ha pagado en la catorcena correspondiente
            aguinaldo_contrato.save()
        """ 
        
        #realiza el calculo del ISR
        calcular_aguinaldo_eventual = 0
        calculo_isr = calcular_isr(request,salario,prima_dominical,calulo_aguinaldo,calcular_aguinaldo_eventual)
            
        #calculo de la prenomina - regla de tres   
        dias_de_pago = 12
        #print("incidencias", incidencias, "incidencias_retarods", incidencias_retardos, "incidencias_inca", incidencias_incapacidades)
        incapacidades = incapacidad_riesgo_laboral + incapacidad_maternidad  + incapacidad_enfermedad_general 
        print("incidencias", incidencias, "incidencias_retarods", incidencias_retardos, "Incapacidades_riesgo", incapacidad_riesgo_laboral + incapacidad_maternidad, + incapacidad_enfermedad_general)
        #se hace el calculo en relacion con los pagos de dias laborados
        #print("estos son los dias a pagar: ", pagar_dias_incapacidad)
        dias_laborados = dias_de_pago - (incidencias + incidencias_retardos + incapacidad_riesgo_laboral + incapacidad_maternidad + (incapacidad_enfermedad_general)) #- pagar_dias_incapacidad))
        print("estos son los dias laborados: ", dias_laborados)
        proporcion_septimos_dias = Decimal((dias_laborados * 2) / 12)
        proporcion_laborados = proporcion_septimos_dias + dias_laborados
        salario_catorcenal = (proporcion_laborados * salario) + pago_doble
        print("ESTE ES EL SALARIO CATORCENAL ", salario_catorcenal)
        
        
        print("ESTE ES EL APOYO PASAJES ahora: ", apoyo_pasajes)
        apoyo_pasajes = (apoyo_pasajes / 12 ) * (12 - (incidencias + incapacidad_enfermedad_general + incapacidad_riesgo_laboral))
          
        print("apoyos pasajes: ", apoyo_pasajes)
        print("total: ", salario_catorcenal)
        print("pagar nomina: ", apoyo_pasajes + salario_catorcenal)
        #total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + prima_dominical
        total_percepciones = salario_catorcenal + apoyo_pasajes + total_bonos + calulo_aguinaldo + prima_dominical +prima_dominical #+ calculo_aguinaldo_eventual
        #IMSS y el ISR
        total_deducciones = prestamo_infonavit + prestamo_fonacot + calculo_isr + calculo_imss
        pagar_nomina = (total_percepciones - total_deducciones)
        
        if retardos == 0: 
            retardos = ''
        
        if castigos == 0:
            castigos = ''
            
        if permisos_con_goce  == 0:
            permisos_con_goce  = ''
            
        if permisos_sin_goce  == 0:
            permisos_sin_goce  = ''
            
        if descansos == 0:
            descansos = ''

        if dia_extra == 0:
            dia_extra = ''
                    
        if incapacidades == 0:
            incapacidades = ''
        
        if faltas == 0:
            faltas = ''
        
        if comisiones == 0:
            comisiones = ''
            
        if domingos == 0:
            domingos = ''
            
        if festivos == 0:
            festivos = ''
            
        if economicos == 0:
            economicos  = ''
        
        if vacaciones == 0:
            vacaciones = ''
        
        # Agregar los valores a la lista rows para cada prenomina
        row = (
            prenomina.empleado.status.perfil.nombres + ' ' + prenomina.empleado.status.perfil.apellidos,
            prenomina.empleado.status.perfil.numero_de_trabajador,
            prenomina.empleado.status.perfil.distrito.distrito,
            catorcena_num,
            str(prenomina.catorcena.fecha_inicial) + " " + str(prenomina.catorcena.fecha_final),
            prenomina.estado_general,
            str(RH),
            str(CT),
            str(G),
            estado,
            retardos,
            castigos,
            permisos_con_goce,
            permisos_sin_goce,
            descansos,
            #str("Días anteriores: ")+str(incapacidades_anterior)+str(" Días actual: ")+str(incapacidades_actual),
            incapacidades,
            faltas,
            comisiones,
            domingos,
            dia_extra,
            festivos,
            economicos,
            vacaciones,
            salario_catorcenal,
            apoyo_pasajes,
            total_bonos,
            total_percepciones,
            prestamo_infonavit,
            calculo_imss,
            prestamo_fonacot,
            calculo_isr,
            #
            #
            total_deducciones,
            pagar_nomina,
        )
        rows.append(row)
        
        #sub_salario_catorcenal_costo = sub_salario_catorcenal_costo + salario_catorcenal_costo
        sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
        sub_apoyo_pasajes = sub_apoyo_pasajes + apoyo_pasajes
        sub_total_bonos = sub_total_bonos + total_bonos
        sub_total_percepciones = sub_total_percepciones + total_percepciones
        sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
        sub_calculo_imss = sub_calculo_imss + calculo_imss
        sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
        sub_calculo_isr = sub_calculo_isr + calculo_isr
        sub_total_deducciones = sub_total_deducciones + total_deducciones
        sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
        
        
                 
    # Ahora puedes usar la lista rows como lo estás haciendo actualmente en tu código
    for row_num, row in enumerate(rows, start=2):
        for col_num, value in enumerate(row, start=1):
            if col_num < 4:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num == 5:
                ws.cell(row=row_num, column=col_num, value=value).style = date_style
            elif col_num > 5 and col_num < 24:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style
            elif col_num >= 24:
                ws.cell(row=row_num, column=col_num, value=value).style = money_style
            else:
                ws.cell(row=row_num, column=col_num, value=value).style = body_style

    #sumar las columnas
    #print("Salario neto cartocelnal", salario_catorcenal)
    #sub_salario_catorcenal = sub_salario_catorcenal + salario_catorcenal
    #print("subtotal catorcenal",sub_salario_catorcenal)
    #sub_apoyo_pasajes = Decimal(sub_apoyo_pasajes) + apoyo_pasajes
    #sub_total_bonos = sub_total_bonos + total_bonos
    #sub_total_percepciones = sub_total_percepciones + total_percepciones
    #sub_prestamo_infonavit = sub_prestamo_infonavit + prestamo_infonavit
    #sub_prestamo_fonacot = sub_prestamo_fonacot + prestamo_fonacot
    #sub_total_deducciones = sub_total_deducciones + total_deducciones
    #sub_pagar_nomina = sub_pagar_nomina + pagar_nomina
    
    
    add_last_row = ['Total','','','','','','','','','','','','','','','','','','','','','','',
                    #sub_salario_catorcenal_costo,
                    sub_salario_catorcenal,
                    sub_apoyo_pasajes,
                    sub_total_bonos,
                    sub_total_percepciones,
                    sub_prestamo_infonavit,
                    sub_calculo_imss,
                    sub_prestamo_fonacot,
                    sub_calculo_isr,
                    sub_total_deducciones,
                    sub_pagar_nomina
                    ]
    ws.append(add_last_row) 
    
    # Aplicar el estilo money_style a cada celda de la fila
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
        for cell in row:
            cell.style = bold_money_style
            
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
        
    return(response)
     