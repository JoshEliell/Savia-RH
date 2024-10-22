from django.shortcuts import render

from django.shortcuts import render,redirect
#verificar autenticacion del usuario
from django.contrib.auth.decorators import login_required
#se importa los modelos de la otra app
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.core.serializers import serialize
import json
from datetime import datetime
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Sum
import os
import logging
from proyecto.models import Distrito,Perfil,Puesto,UserDatos,Catorcenas,DatosBancarios
from .models import Categoria,Subcategoria,Bono,Solicitud,BonoSolicitado,Requerimiento
from revisar.models import AutorizarSolicitudes
from .forms import SolicitudForm,BonoSolicitadoForm,RequerimientoForm,AutorizarSolicitudesUpdateForm,AutorizarSolicitudesGerenteUpdateForm,BonoSolicitadoPuestoForm,AutorizarSolicitudForm
from django.db import connection
from django.core.paginator import Paginator
from .filters import SolicitudFilter,AutorizarSolicitudesFilter,BonoSolicitadoFilter
from django.db.models import Max
from django.db.models import Subquery, OuterRef
from django.http import Http404
import datetime
from datetime import date, timedelta
from django.db.models import Q
#Excel
from openpyxl import Workbook
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.db.models.functions import Concat
from django.db.models import Value
from django.db.models import Sum
from django.db.models import Count
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.http import HttpResponseRedirect
from datetime import datetime
from datetime import datetime, timedelta
from django.utils import timezone
from datetime import date
#pillow - comprimir imagenes
from PIL import Image
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
#envio de correos
from django.core.mail import send_mail
from django.conf import settings
from decimal import Decimal
# PyMuPDF - comprimir pdfs
import fitz  # PyMuPDF
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import io
from user.decorators import perfil_session_seleccionado


#Pagina inicial de los esquemas de los bonos
@login_required(login_url='user-login')
@perfil_session_seleccionado
def inicio(request):
    bonos = Categoria.objects.all();
    context= {
        'bonos':bonos,
    }
    return render(request,'esquema/inicio.html',context)

#Listar las solicitudes
@login_required(login_url='user-login')
@perfil_session_seleccionado
def listarBonosVarilleros(request):    
    ids = [9,10,11]    
    
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    if usuario.tipo not in [1,2,3]:
                
        subconsulta_ultima_fecha = AutorizarSolicitudes.objects.values('solicitud_id').annotate(
                ultima_fecha=Max('created_at')
            ).filter(solicitud_id=OuterRef('solicitud_id')).values('ultima_fecha')[:1]
        
        if usuario.tipo.id in [9,10,11]:
            #obtiene todas las ultimas autorizaciones de todos los distritos y roles independientemente en el flujo que se encuentre
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__complete = 1
            ).order_by("-created_at")
        elif usuario.tipo.id in [4,5]: #rh, supervisor - Ve todas sus solicitudes creadas y en el flujo en que se encuentran
            #obtiene todas las ultimas autorizaciones de su distrito y roles
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__solicitante__id = usuario.perfil_id, solicitud__distrito_id = usuario.distrito.id,solicitud__complete = 1, #mostrar solamente sus solicitudes y de nadie mas
            ).filter(
                Q(estado_id = 3) | Q(estado_id = 4)
            ).order_by("-created_at")
        elif usuario.tipo.id in [8]: #GE puede ver todas las solicitudes creadas y en el flujo en el que se encuentran
            #obtiene la ultima autorizacion independientemente en el flujo que se encuentre     
            print("se ejecuta GE")       
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__distrito_id = usuario.distrito.id, solicitud__complete = 1, estado_id = 3,tipo_perfil_id = usuario.tipo.id
            ).order_by("-created_at")
        else:
            #solo obtiene la solicitud que le pertenece    
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                perfil_id = usuario.perfil.id, tipo_perfil_id = usuario.tipo.id, solicitud__distrito_id = usuario.distrito.id, solicitud__complete = 1, estado_id = 3
            ).order_by("-created_at")
        
        autorizaciones_filter = AutorizarSolicitudesFilter(request.GET, queryset=autorizaciones)
        autorizaciones = autorizaciones_filter.qs
            
        p = Paginator(autorizaciones, 50)
        page = request.GET.get('page')
        salidas_list = p.get_page(page)
        autorizaciones= p.get_page(page)
        
        contexto = {
            #'usuario':usuario,
            'autorizaciones':autorizaciones,
            'autorizaciones_filter': autorizaciones_filter,
            'salidas_list':salidas_list,
            'ids': ids
        }
        
        return render(request,'esquema/bonos_varilleros/listar.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')
    
def comprimir_imagen(imagen):
    """
    Comprime una imagen y devuelve un objeto InMemoryUploadedFile.
    
    Args:
    imagen (InMemoryUploadedFile): Objeto de imagen subida.
    
    Returns:
    InMemoryUploadedFile: Objeto de imagen comprimida.
    """
    # Abre la imagen usando Pillow
    img = Image.open(imagen)
    
    if img.format != 'JPEG':
        img = img.convert('RGB')
    
    # Crea un flujo de bytes para almacenar la imagen comprimida
    img_temp_output = BytesIO()
    
    ancho_original, alto_original = img.size

    # Hace que el height y width se reduzca a la mitad
    nuevo_ancho = ancho_original // 2
    nuevo_alto = alto_original // 2

    # Redimensionar la imagen
    img = img.resize((nuevo_ancho, nuevo_alto))
    
    # Comprime la imagen y la guarda en el flujo de bytes
    img.save(img_temp_output, format='JPEG', quality=50, optimize=True)  # Puedes ajustar la calidad según tus necesidades
    
    # Restablece el puntero del flujo de bytes al principio
    img_temp_output.seek(0)
    
    # Crea un objeto InMemoryUploadedFile para la imagen comprimida
    img_comprimida = InMemoryUploadedFile(img_temp_output, None, imagen.name.split('.')[0] + 'b.jpg', 'image/jpeg', img_temp_output.getbuffer().nbytes, None)
    
    return img_comprimida

def comprimir_pdf(pdf):
    # Open the PDF using PyMuPDF
    pdf_document = fitz.open(stream=pdf.read(), filetype='pdf')
    
    # Create an in-memory buffer for the compressed PDF
    output_buffer = io.BytesIO()
    
    # Define compression settings (for images)
    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        images = page.get_images(full=True)
        
        # Dictionary to store the new image data
        new_images = {}
        
        for img_index, img in enumerate(images):
            xref = img[0]
            base_image = pdf_document.extract_image(xref)
            image_bytes = base_image['image']
            
            # Convert image bytes to a Pillow Image object
            with Image.open(io.BytesIO(image_bytes)) as image:
                # Convert to RGB (in case it's RGBA or another mode)
                image = image.convert("RGB")
                
                #Redimensionar la imagen
                image = image.resize((image.width // 2, image.height // 2), Image.LANCZOS)  # Use Image.LANCZOS for quality reduction
                
                # Create a new in-memory buffer for the compressed image
                img_buffer = io.BytesIO()
                
                # Save the image with reduced quality using Pillow
                image.save(img_buffer, format='JPEG', quality=80, optimize=True)  # Adjust quality as needed
                
                # Store the new image data
                new_images[xref] = img_buffer.getvalue()
        
        # Remove old images and insert new ones
        for xref, new_img_data in new_images.items():
            # Remove the image
            page.delete_image(xref)
            # Insert the new image
            page.insert_image(page.rect, stream=new_img_data)
    
    # Save the compressed PDF to the buffer
    pdf_document.save(output_buffer)
    pdf_document.close()
    
    # Rewind the buffer to the beginning
    output_buffer.seek(0)
    
    # Obtener la fecha y hora actual
    now = datetime.now()
    # Formatear la fecha como una cadena (por ejemplo, "2024-08-09_15-30-45")
    fecha_formateada = now.strftime("%Y-%m-%d_%H-%M-%S")
    # Crear el nombre del archivo usando la fecha
    nombre_archivo = f"file_{fecha_formateada}.pdf"

    # Create an InMemoryUploadedFile object
    compressed_pdf_file = InMemoryUploadedFile(
        output_buffer,
        None,
        nombre_archivo,
        'application/pdf',
        output_buffer.getbuffer().nbytes,
        None
    )
    
    # Return the buffer with the compressed PDF
    return compressed_pdf_file
    
    
#para crear solicitudes de bonos
@login_required(login_url='user-login')
@perfil_session_seleccionado
def crearSolicitudBonos(request):
    
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id) 
    
    #Todos los supervisores y RH pueden crear solicitudes
    if usuario.tipo_id in (5,4): 
        bonoSolicitadoForm = BonoSolicitadoForm()
        bonoSolicitadoPuestoForm = BonoSolicitadoPuestoForm()
        requerimientoForm = RequerimientoForm()
        autorizarSolicitudForm = AutorizarSolicitudForm()
        #se hace una consulta con los empleados del distrito que pertenecen
        empleados = Perfil.objects.filter(distrito_id = usuario.distrito.id).exclude(baja = 1).order_by('nombres')        
        solicitante = usuario.perfil
        #obtener los superintendentes operativos, administrativos
        perfiles = UserDatos.objects.select_related('perfil').filter(distrito_id = usuario.distrito.id, activo=True,tipo_id__in=[6,12])
        perfiles_id = [p.perfil.id for p in perfiles]
        
        bonoSolicitadoForm.fields["trabajador"].queryset = empleados 
        folio = Solicitud.objects.filter(complete=True).order_by('-folio').values_list('folio', flat=True).first() + 1
        solicitud, created = Solicitud.objects.get_or_create(complete = False, defaults={'complete': False, 'folio':folio,'solicitante_id':solicitante.id, 'total':0.00})
        #Obtiene los bonos que han sido creados en la solicitud
        bonos_solicitados = BonoSolicitado.objects.filter(solicitud_id = solicitud.id)
        lista_archivos = Requerimiento.objects.filter(solicitud_id = solicitud.id)
        
        bonos = Subcategoria.objects.all()
        
        bonos_para_select2 = [
            {
                'id': bono.id,
                'text': str(bono.nombre),
                'soporte': str(bono.soporte)
            } for bono in bonos
        ]
        
        errors = {}
        
        #Para guardar la solicitud
        if request.method == "POST":
            #obtiene un queryset de los archivos de la solicitud
        
            if 'btn_archivos' in request.POST:      
                #Se envian los formularios con datos                   
                form = RequerimientoForm(request.POST, request.FILES)  
                archivos = request.FILES.getlist('url')
                    
                if form.is_valid():
                    #Se recorren los archivos para ser almacenados
                    for archivo in archivos:
                        #obtener el tam del archivo cargado en megas
                        archivo_tam = archivo.size/(1024 * 1024)
                        if archivo.content_type != 'application/pdf' and archivo.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and archivo.content_type != 'application/vnd.ms-excel':
                            if archivo_tam > 5: # 5 megas
                                documento = comprimir_imagen(archivo)
                            else:
                                documento = archivo
                        #Comprime pdf
                        elif archivo.content_type == 'application/pdf':
                            if archivo_tam > 8: # 8 megas
                                documento = comprimir_pdf(archivo)
                            else:
                                documento = archivo
                        else:
                            documento = archivo
                        #Guarda el archivo
                        requerimiento = Requerimiento.objects.create(
                            solicitud_id = solicitud.id,
                            url = documento,
                        )
                        
                        requerimiento.save()
                        
                    solicitud.complete_requerimiento = True
                    solicitud.save()
                    messages.success(request, "El soporte se adjunto correctamente")
                    return redirect("crearSolicitudBonos") 
                else:
                    # Mensaje de error si el formulario no es válido
                    messages.error(request, 'Favor de verificar el formato del soporte y el tamaño maximo son 10 MB')
                
            elif 'btn_agregar' in request.POST:
                form_solicitud = SolicitudForm(request.POST, instance = solicitud)
                if form_solicitud.is_valid():
                    esquema_bono = request.POST.get('esquemaBono')
                    bono_solicitado, created = BonoSolicitado.objects.get_or_create(solicitud_id=solicitud.id,bono_id = esquema_bono,cantidad = 0)
                    form_bonosolicitado = BonoSolicitadoForm(request.POST,instance=bono_solicitado)
                    if form_bonosolicitado.is_valid():
                        puesto = int(request.POST.get('puesto'))
                        if puesto == 7: # Puesto ID - Todos los que participen en la actividad
                            bono_solicitado.save()
                            # id puesto - todos los que participen en la actividad - dividir la cantidad del bono entre el total de los participantes
                            participantes = BonoSolicitado.objects.filter(solicitud_id = solicitud.id).count()
                            #División
                            reparto = Decimal(bono_solicitado.cantidad/participantes)
                            solicitud.total = bono_solicitado.cantidad
                            solicitud.save()
                            BonoSolicitado.objects.filter(solicitud_id = solicitud.id).update(cantidad=reparto)
                        
                        elif bono_solicitado.bono.esquema_subcategoria.id == 21: # Bono ó Subcategoria ID - TOMA DE REGISTROS ECO Y/O DINA - 9 soportes se paga, 30% ayudante, 70% tecnico
                            servicios = Requerimiento.objects.filter(solicitud_id = solicitud.id,url__iendswith='.pdf').count() # solo se cuenta el numero de pdfs
                            if servicios >= 9: #9 valor, Paga el novevo servicio - soporte
                                bono_solicitado.save()
                                cantidad = bono_solicitado.cantidad * (servicios - 8) #8 valor
                                puestos = BonoSolicitado.objects.filter(solicitud_id = solicitud.id)
                                if puestos.count() > 1:
                                    BonoSolicitado.objects.filter(
                                        solicitud_id=solicitud.id,
                                        bono__puesto_id=16 # TÉCNICO DE TOMA DE INFORMACIÓN
                                    ).update(cantidad=Decimal('0.70') * Decimal(cantidad))
                                    BonoSolicitado.objects.filter(
                                        solicitud_id=solicitud.id,
                                        bono__puesto_id=39  # AYUDANTE
                                    ).update(cantidad=Decimal('0.30') * Decimal(cantidad))
                                    solicitud.total = Decimal(cantidad)
                                    solicitud.save()
                                else:
                                    BonoSolicitado.objects.filter(solicitud_id=solicitud.id).update(cantidad=cantidad)
                                    solicitud.total = Decimal(cantidad)
                                    solicitud.save()
                            else:
                                bono_solicitado.cantidad = Decimal(0.00)
                                bono_solicitado.save()
                                solicitud.total = Decimal(0.00)
                                solicitud.save()
                        
                        else: # EN CASO DE OTRO BONO
                            bono_solicitado.bono.id = esquema_bono
                            bono_solicitado.save()
                            total = BonoSolicitado.objects.filter(solicitud_id = solicitud.id).aggregate(total=Sum('cantidad'))['total'] 
                            solicitud.total = total
                            solicitud.save()
                            
            elif 'enviar_solicitud' in request.POST:
                #validacion para subir los soportes
                if solicitud.complete_requerimiento == 0:
                    messages.error(request, 'Falta adjuntar soportes')
                    return redirect(request.META.get('HTTP_REFERER'))
                
                autorizarSolicitudForm = AutorizarSolicitudForm(request.POST)
                
                if autorizarSolicitudForm.is_valid():    
                    autorizar = autorizarSolicitudForm.save(commit=False)  
                    
                    #se obtiene el rol del perifil
                    superintendente = UserDatos.objects.get(perfil_id = autorizar.perfil.id,distrito_id = usuario.distrito.id, tipo_id__in=[6,12],activo = True)

                    #guardar en la BD
                    autorizar.solicitud_id = solicitud.id
                    autorizar.tipo_perfil_id= superintendente.tipo_id
                    autorizar.estado_id = 3
                    autorizar.tipo_id = superintendente.tipo.id
                    autorizar.save()
                    #Se guarda la solicitud en complete
                    solicitud.distrito_id = usuario.distrito.id
                    solicitud.complete = 1
                    solicitud.save()
                    messages.success(request, f"La solicitud se envio a {autorizar.perfil}")
                    return redirect('listarBonosVarilleros')
                
        solicitudForm =  SolicitudForm(instance = solicitud)
        autorizarSolicitudForm.fields['perfil'].queryset = Perfil.objects.filter(id__in=perfiles_id)
    
        contexto = {
            'folio': folio,
            'bonos': bonos_para_select2,
            'usuario':usuario,
            'solicitante':solicitante,
            'solicitudForm':solicitudForm,
            'bonoSolicitadoForm':bonoSolicitadoForm,
            'requerimientoForm':requerimientoForm,
            'autorizarSolicitudForm':autorizarSolicitudForm,
            'bonoSolicitadoPuestoForm':bonoSolicitadoPuestoForm,
            'lista_archivos': lista_archivos,
            'bonos_solicitados': bonos_solicitados,
            'errors':errors,
            'solicitud':solicitud,

        } 

        return render(request,'esquema/bonos_varilleros/crear_solicitud.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')
        
    

@login_required(login_url='user-login')  
@perfil_session_seleccionado
def verificarSolicitudBonosVarilleros(request,solicitud):
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    #solamente RH y supervisores
    if usuario.tipo.id in (4,5):
        
        #perfil = Perfil.objects.filter(numero_de_trabajador=usuario.numero_de_trabajador,distrito_id = usuario.distrito.id).values_list('id',flat=True)
        permiso = Solicitud.objects.filter(solicitante_id=usuario.perfil.id, pk=solicitud).values_list('id',flat=True)
        
        #checa el perfil correspondiente para cambiar la solicitud - policy
        if not permiso:
            return render(request, 'revisar/403.html')
        
        #obtener las instancias para poblar los formularios
        solicitud = Solicitud.objects.get(pk=solicitud)
        solicitudForm =  SolicitudForm(instance = solicitud)
        
        #se llaman los formularios para su posterior llenado
        bonoSolicitadoForm = BonoSolicitadoForm()
        bonoSolicitadoPuestoForm = BonoSolicitadoPuestoForm()
        requerimientoForm = RequerimientoForm()
        
        #se hace una consulta con los empleados del distrito que pertenecen
        empleados = Perfil.objects.filter(distrito_id = usuario.distrito.id).exclude(baja = 1).order_by('nombres')
        solicitante = Perfil.objects.get(pk = usuario.perfil.id, distrito_id = usuario.distrito.id)
        #se carga el formulario en automatico definiendo filtros
        bonoSolicitadoForm.fields["trabajador"].queryset = empleados 
        #se llama la autorizacion relacionada
        comentarios = AutorizarSolicitudes.objects.filter(solicitud_id=solicitud.id).exclude(comentario__isnull=True).values_list('comentario', flat=True).order_by('-id')
        
        autorizacion = AutorizarSolicitudes.objects.filter(solicitud_id=solicitud.id).first()
        
        bonos_solicitados = BonoSolicitado.objects.filter(solicitud_id = solicitud.id)
        lista_archivos = Requerimiento.objects.filter(solicitud_id = solicitud.id)
        
        #Se obtiene los nombres de los bonos
        bonos = Subcategoria.objects.all()
        
        bonos_para_select2 = [
            {
                'id': bono.id,
                'text': str(bono.nombre),
                'soporte': str(bono.soporte)
            } for bono in bonos
        ]
        
        errors = {}
        #Para guardar la solicitud
        if request.method == "POST":
            #obtiene un queryset de los archivos de la solicitud

            if 'btn_archivos' in request.POST:      
                #Se envian los formularios con datos                   
                form = RequerimientoForm(request.POST, request.FILES)  
                archivos = request.FILES.getlist('url')
                
                if form.is_valid():
                    #Se recorren los archivos para ser almacenados
                    for archivo in archivos:
                        #obtener el tam del archivo cargado en megas
                        archivo_tam = archivo.size/(1024 * 1024)
                        if archivo.content_type != 'application/pdf' and archivo.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and archivo.content_type != 'application/vnd.ms-excel':
                            if archivo_tam > 5: # 5 megas
                                documento = comprimir_imagen(archivo)
                            else:
                                documento = archivo
                        #Comprime pdf
                        elif archivo.content_type == 'application/pdf':
                            if archivo_tam > 8: # 8 megas
                                documento = comprimir_pdf(archivo)
                            else:
                                documento = archivo
                        else:
                            documento = archivo
                        #Guarda el archivo
                        requerimiento = Requerimiento.objects.create(
                            solicitud_id = solicitud.id,
                            url = documento,
                        )
                        
                        requerimiento.save()
                        
                    solicitud.complete_requerimiento = True
                    solicitud.save()
                    messages.success(request, "El soporte se adjunto correctamente")
                    return redirect('verificarSolicitudBonosVarilleros', solicitud=solicitud.id) 
                else:
                    # Mensaje de error si el formulario no es válido
                    messages.error(request, 'Favor de verificar el formato del soporte y el tamaño maximo son 10 MB')
                                
            elif 'btn_agregar' in request.POST:
                form_solicitud = SolicitudForm(request.POST, instance = solicitud)
                if form_solicitud.is_valid():
                    esquema_bono = request.POST.get('esquemaBono')
                    bono_solicitado, created = BonoSolicitado.objects.get_or_create(solicitud_id=solicitud.id,bono_id = esquema_bono,cantidad = 0)
                    form_bonosolicitado = BonoSolicitadoForm(request.POST,instance=bono_solicitado)
                    if form_bonosolicitado.is_valid():
                        puesto = int(request.POST.get('puesto'))
                        if puesto == 7: #
                            bono_solicitado.save()
                            # id puesto - todos los que participen en la actividad - dividir la cantidad del bono entre el total de los participantes
                            participantes = BonoSolicitado.objects.filter(solicitud_id = solicitud.id).count()
                            #División
                            reparto = Decimal(bono_solicitado.cantidad/participantes)
                            solicitud.total = bono_solicitado.cantidad
                            solicitud.save()
                            BonoSolicitado.objects.filter(solicitud_id = solicitud.id).update(cantidad=reparto)
                        
                        elif bono_solicitado.bono.esquema_subcategoria.id == 21: # Bono ó Subcategoria ID - TOMA DE REGISTROS ECO Y/O DINA - 9 soportes se paga, 30% ayudante, 70% tecnico
                            servicios = Requerimiento.objects.filter(solicitud_id = solicitud.id,url__iendswith='.pdf').count()
                            if servicios >= 9: #Paga el novevo servicio - soporte
                                bono_solicitado.save()
                                cantidad = bono_solicitado.cantidad * (servicios - 8)
                                puestos = BonoSolicitado.objects.filter(solicitud_id = solicitud.id)
                                if puestos.count() > 1:
                                    BonoSolicitado.objects.filter(
                                        solicitud_id=solicitud.id,
                                        bono__puesto_id=16 # TÉCNICO DE TOMA DE INFORMACIÓN
                                    ).update(cantidad=Decimal('0.70') * Decimal(cantidad))
                                    BonoSolicitado.objects.filter(
                                        solicitud_id=solicitud.id,
                                        bono__puesto_id=39  # AYUDANTE
                                    ).update(cantidad=Decimal('0.30') * Decimal(cantidad))
                                    solicitud.total = Decimal(cantidad)
                                    solicitud.save()
                                else:
                                    BonoSolicitado.objects.filter(solicitud_id=solicitud.id).update(cantidad=cantidad)
                                    solicitud.total = Decimal(cantidad)
                                    solicitud.save()
                            else:
                                bono_solicitado.cantidad = Decimal(0.00)
                                bono_solicitado.save()
                                solicitud.total = Decimal(0.00)
                                solicitud.save()
                                
                        else:
                            bono_solicitado.bono.id = esquema_bono
                            bono_solicitado.save()
                            total = BonoSolicitado.objects.filter(solicitud_id = solicitud.id).aggregate(total=Sum('cantidad'))['total'] 
                            solicitud.total = total
                            solicitud.save()
                            
            elif 'actualizar_solicitud' in request.POST:    
                autorizacion.estado_id = 3 # Estado pendiente
                autorizacion.save()
                messages.success(request, "La solicitud se envio con nuevos cambios al superintendente")
                return redirect('listarBonosVarilleros')
                
        solicitudForm = SolicitudForm(instance = solicitud)
    
        contexto = {
           
            'bonos': bonos_para_select2,
            'usuario':usuario,
            'solicitante':solicitante,
            'solicitudForm':solicitudForm,
            'bonoSolicitadoForm':bonoSolicitadoForm,
            'requerimientoForm':requerimientoForm,
            'bonoSolicitadoPuestoForm':bonoSolicitadoPuestoForm,
            'lista_archivos': lista_archivos,
            'bonos_solicitados': bonos_solicitados,
            'errors':errors,
            'solicitud':solicitud,
            'autorizacion':autorizacion,
            'comentarios': comentarios
        
        } 
        
        return render(request, 'esquema/bonos_varilleros/verificar_solicitud.html', contexto)
    
    else:
        return render(request, 'revisar/403.html')
        
#para ver detalles de la solicitud cuando se debe realizar la autorizacion 
@login_required(login_url='user-login')
@perfil_session_seleccionado
def verDetallesSolicitud(request,solicitud_id):  
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    if usuario.tipo not in [1,2,3]:
        #busca la ultima solicitud con relacion a sus modelos     
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud_id=solicitud_id
        ).select_related(
            'solicitud',  # Relación con Solicitud
            'solicitud__solicitante',  # Relación a Perfil desde Solicitud
            'solicitud__bono',  # Relación a Subcategoria desde Solicitud
            'solicitud__distrito',  # Relación a Distrito desde Solicitud
        ).prefetch_related(
            'solicitud__bonos_solicitados',  # Relación a BonoSolicitado desde Solicitud (usa related_name si está definido)
            'solicitud__requerimientos',  # Relación a Requerimiento desde Solicitud (usa related_name si está definido)
        ).annotate(
            ultima_fecha=Max('created_at'),
        ).order_by('-ultima_fecha').first()
        
        
        #Se cuenta el numero de soportes y el tipo de extension
        archivos_pdf = autorizaciones.solicitud.requerimientos.filter(url__iendswith='.pdf').count()
        archivos_excel = autorizaciones.solicitud.requerimientos.filter(Q(url__iendswith='.xls') | Q(url__iendswith='.xlsx')).count()
        archivos_imagenes = autorizaciones.solicitud.requerimientos.filter(Q(url__iendswith='.png') | Q(url__iendswith='.jpeg') | Q(url__iendswith='.jpeg')).count()
        total_archivos = archivos_pdf + archivos_excel + archivos_imagenes
             
        if usuario.tipo.id not in [9,10,11] and usuario.distrito.id != autorizaciones.solicitud.solicitante.distrito.id:
            return render(request, 'revisar/403.html')
                
        #obtener el rol del solicitante    
        rol = UserDatos.objects.get(perfil_id = autorizaciones.solicitud.solicitante.id, distrito_id = autorizaciones.solicitud.distrito.id, tipo_id__in=[4,5], activo = True)
        
        #se carga el formulario con datos iniciales
        autorizarSolicitudesUpdateForm = AutorizarSolicitudesUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
        autorizarSolicitudesGerenteUpdateForm = AutorizarSolicitudesGerenteUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
            
        contexto = {
            "usuario":usuario,
            "autorizaciones":autorizaciones,
            "autorizarSolicitudesUpdateForm":autorizarSolicitudesUpdateForm,
            "autorizarSolicitudesGerenteUpdateForm":autorizarSolicitudesGerenteUpdateForm,
            "rol":rol,
            "archivos_pdf":archivos_pdf,
            "archivos_excel":archivos_excel,
            "archivos_imagenes":archivos_imagenes,
            "total_archivos":total_archivos
        }
        
        return render(request,'esquema/bonos_varilleros/detalles_solicitud.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')

@login_required(login_url='user-login')
@perfil_session_seleccionado
def detalleSolicitudAutorizacion(request,solicitud_id):
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    if usuario.tipo not in [1,2,3]:
        
        #busca la ultima solicitud con relacion a sus modelos     
        autorizaciones = AutorizarSolicitudes.objects.filter(
            #solicitud_id=solicitud_id, perfil_id = usuario.perfil.id, tipo_perfil_id = usuario.tipo.id
            solicitud_id=solicitud_id
        ).select_related(
            'solicitud',  # Relación con Solicitud
            'solicitud__solicitante',  # Relación a Perfil desde Solicitud
            'solicitud__bono',  # Relación a Subcategoria desde Solicitud
            'solicitud__distrito',  # Relación a Distrito desde Solicitud
        ).prefetch_related(
            'solicitud__bonos_solicitados',  # Relación a BonoSolicitado desde Solicitud (usa related_name si está definido)
            'solicitud__requerimientos',  # Relación a Requerimiento desde Solicitud (usa related_name si está definido)
        ).order_by()
         
        flujo_autorizaciones = autorizaciones
        autorizaciones = autorizaciones.first()
        
        #Se cuenta el numero de soportes y el tipo de extension
        archivos_pdf = autorizaciones.solicitud.requerimientos.filter(url__iendswith='.pdf').count()
        archivos_excel = autorizaciones.solicitud.requerimientos.filter(Q(url__iendswith='.xls') | Q(url__iendswith='.xlsx')).count()
        archivos_imagenes = autorizaciones.solicitud.requerimientos.filter(Q(url__iendswith='.png') | Q(url__iendswith='.jpeg') | Q(url__iendswith='.jpeg')).count()
        total_archivos = archivos_pdf + archivos_excel + archivos_imagenes
        
        #Permiso - solo el usuario perteneciente al distrito puede ver los bonos
        if usuario.tipo.id not in [9,10,11] and usuario.distrito.id != autorizaciones.solicitud.distrito.id:
            return render(request, 'revisar/403.html')
                
        #obtener el rol del solicitante    
        rol = UserDatos.objects.get(perfil_id = autorizaciones.solicitud.solicitante.id, distrito_id = autorizaciones.solicitud.distrito.id, tipo_id__in=[4,5], activo = True)
        
        #se carga el formulario con datos iniciales
        autorizarSolicitudesUpdateForm = AutorizarSolicitudesUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
        autorizarSolicitudesGerenteUpdateForm = AutorizarSolicitudesGerenteUpdateForm(initial={'estado':autorizaciones.estado.id,'comentario':autorizaciones.comentario})
        
        contexto = {
            "usuario":usuario,
            "autorizaciones":autorizaciones,
            "autorizarSolicitudesUpdateForm":autorizarSolicitudesUpdateForm,
            "autorizarSolicitudesGerenteUpdateForm":autorizarSolicitudesGerenteUpdateForm,
            "rol":rol,
            "flujo_autorizaciones":flujo_autorizaciones,
            "archivos_pdf":archivos_pdf,
            "archivos_excel":archivos_excel,
            "archivos_imagenes":archivos_imagenes,
            "total_archivos":total_archivos,
        }
        
        return render(request,'esquema/bonos_varilleros/detalle_solicitud_autorizada.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')    
    

#Listar todas las solicitudes de los bonos
@login_required(login_url='user-login')
@perfil_session_seleccionado
def todasSolicitudesBonos(request):
    ids = [9,10,11]   
    
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    if usuario.tipo not in [1,2,3]:
          
        subconsulta_ultima_fecha = AutorizarSolicitudes.objects.values('solicitud_id').annotate(
                ultima_fecha=Max('created_at')
            ).filter(solicitud_id=OuterRef('solicitud_id')).values('ultima_fecha')[:1]
        
        if usuario.tipo.id in [9,10,11]:
            #obtiene todas las ultimas autorizaciones de todos los distritos y roles independientemente en el flujo que se encuentre
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__complete = 1
            ).order_by("-created_at")
        elif usuario.tipo.id in [4,5]: #rh, supervisor - Ve todas sus solicitudes creadas y en el flujo en que se encuentran
            #obtiene todas las ultimas autorizaciones de su distrito y roles
            autorizaciones = AutorizarSolicitudes.objects.filter(
                created_at=Subquery(subconsulta_ultima_fecha)
            ).select_related('solicitud', 'perfil').filter(
                solicitud__solicitante__id = usuario.perfil_id, solicitud__distrito_id = usuario.distrito.id,solicitud__complete = 1, #mostrar solamente sus solicitudes y de nadie mas
            ).order_by("-created_at")
        elif usuario.tipo.id in [8]: #GE puede ver todas las solicitudes creadas y en el flujo en el que se encuentran
            #obtiene la ultima autorizacion independientemente en el flujo que se encuentre           
            autorizaciones = AutorizarSolicitudes.objects.select_related('solicitud', 'perfil').filter(
                solicitud__distrito_id = usuario.distrito.id, solicitud__complete = 1,tipo_perfil_id = usuario.tipo.id
            ).order_by("-created_at")
        else:
            #solo obtiene la solicitud que le pertenece    
            autorizaciones = AutorizarSolicitudes.objects.select_related('solicitud', 'perfil').filter(
                perfil_id = usuario.perfil.id, tipo_perfil_id = usuario.tipo.id, solicitud__distrito_id = usuario.distrito.id, solicitud__complete = 1
            ).order_by("-created_at")
        
        autorizaciones_filter = AutorizarSolicitudesFilter(request.GET, queryset=autorizaciones)
        autorizaciones = autorizaciones_filter.qs
            
        p = Paginator(autorizaciones, 50)
        page = request.GET.get('page')
        salidas_list = p.get_page(page)
        autorizaciones= p.get_page(page)
        
        contexto = {
            #'usuario':usuario,
            'autorizaciones':autorizaciones,
            'autorizaciones_filter': autorizaciones_filter,
            'salidas_list':salidas_list,
            'ids': ids
        }
        
        return render(request,'esquema/bonos_varilleros/todas_solicitudes.html',contexto)
    
    else:
        return render(request, 'revisar/403.html')
    
   

#lista bonos aprobados
@login_required(login_url='user-login')
@perfil_session_seleccionado
def listarBonosVarillerosAprobados(request):
    from django.db.models import Prefetch
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    ids = [9,10,11]
    
    #Se muestran por catorcenas
    fecha_actual = datetime.now()
    fecha_actual = fecha_actual - timedelta(days=25) #ELIMINAR FECHA DE CATORCENA ANTERIOR
    
    catorcena_actual = Catorcenas.objects.filter(fecha_inicial__lte=fecha_actual, fecha_final__gte=fecha_actual).first()
    fecha_inicial = datetime.combine(catorcena_actual.fecha_inicial, datetime.min.time()) + timedelta(hours=00, minutes=00,seconds=00)
    fecha_final = datetime.combine(catorcena_actual.fecha_final, datetime.min.time()) + timedelta(hours=23, minutes=59,seconds=59)

    #flujo permisos y autorizaciones
    autorizaciones = None
    #Si es usuario RH de distrito matriz
    if usuario.tipo.id in (9,10,11):
        #obtiene todos los bonos aprobados de todos los distritos
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud__complete = 1,
            estado_id = 1,
            tipo_perfil_id = 8,
            updated_at__range=(fecha_inicial,fecha_final)
           
        ).order_by("-created_at").values('solicitud_id')
    
    elif usuario.tipo.id in (4,12,8):
        #obtiene todos los bonos aprobados de un solo distrito al que pertenece
        autorizaciones = AutorizarSolicitudes.objects.filter(
            solicitud__complete = 1,
            estado_id = 1,
            tipo_perfil_id = 8,
            perfil__distrito_id = usuario.distrito.id,
            updated_at__range=(fecha_inicial,fecha_final)
           
        ).order_by("-created_at").values('solicitud_id')

        """
        bonos_solicitados = BonoSolicitado.objects.filter(
            solicitud__autorizarsolicitudes__isnull=False,
            solicitud__autorizarsolicitudes__estado_id=1,  # Puedes ajustar los filtros según lo que necesites
            solicitud__autorizarsolicitudes__tipo_perfil_id=8,
            solicitud__autorizarsolicitudes__perfil__distrito_id=usuario.distrito.id,
            solicitud__autorizarsolicitudes__updated_at__range=(fecha_inicial, fecha_final)
        )
        
        for b in bonos_solicitados:
            print(b)
            print(b.trabajador)
            print(b.trabajador.numero_de_trabajador)
            print(b.bono)
            print(b.cantidad)
            print(b.trabajador.status.datosbancarios.no_de_cuenta)
        
        """
        
    else:
        return render(request, 'revisar/403.html')
        
    #Permisos
    if autorizaciones is None:
        return render(request, 'revisar/403.html')
    
    #se buscan los perfiles acredores al bono
    solicitudes = []
    for item in autorizaciones:
        solicitud_id = item['solicitud_id']
        solicitudes.append(solicitud_id)
        
    bonos = BonoSolicitado.objects.filter(solicitud_id__in = solicitudes).order_by('trabajador_id')
    bonosolicitado_filter = BonoSolicitadoFilter(request.GET, queryset=bonos) 
    bonos = bonosolicitado_filter.qs
    
    total_monto = bonos.aggregate(total_monto=Sum('cantidad'))['total_monto']
    cantidad_bonos_aprobados = bonos.count()
    
    if request.method =='POST' and 'excel' in request.POST:
        return convert_excel_bonos_aprobados(bonos,catorcena_actual,total_monto,cantidad_bonos_aprobados)
    
    p = Paginator(bonos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)
    bonos = p.get_page(page)
    
    contexto = {
        'bonos':bonos,
        'salidas_list':salidas_list,
        'bonosolicitado_filter':bonosolicitado_filter,
        'cantidad_bonos_aprobados':cantidad_bonos_aprobados,
        'total_monto':total_monto,
        'catorcena':catorcena_actual,
        'usuario':usuario,
        'ids':ids
    }
    
    return render(request,'esquema/bonos_varilleros/listar_bonos_aprobados.html',contexto)

#generar reportes bonos aprobados
@login_required(login_url='user-login')
@perfil_session_seleccionado
def generarReporteBonosVarillerosAprobados(request):
    #obtener datos de la sesion y el usuario logeado
    userdatos = request.session.get('usuario_datos')        
    usuario_id = userdatos.get('usuario_id')
    usuario = UserDatos.objects.get(pk = usuario_id)
    
    ids = [9,10,11]
        
    #Flujo de las autorizaciones y permisos
    if usuario.tipo.id in (9,10,11,12): #
        #se buscan los perfiles acredores al bono
        folios = Solicitud.objects.filter(fecha_autorizacion__isnull=False).values('id')
    elif usuario.tipo.id in (4,12,8): #RH, SA, GE
        folios = Solicitud.objects.filter(fecha_autorizacion__isnull=False, solicitante__distrito_id = usuario.distrito.id).values('id')
    else:
        return render(request, 'revisar/403.html')
   
    
    #se prepara un 
    solicitudes = []
    for item in folios:
        solicitud_id = item['id']
        solicitudes.append(solicitud_id)
        
    bonos = BonoSolicitado.objects.filter(solicitud_id__in = solicitudes)     
    bonosolicitado_filter = BonoSolicitadoFilter(request.GET, queryset=bonos) 
    bonos = bonosolicitado_filter.qs
    
    bono = bonos.last()
    
    if not request.GET:
        catorcena = None
    else:
        try:#se maneja una expecion cuando no se encuentran bonos en una catorcena o fecha especificada
            catorcena = Catorcenas.objects.filter(fecha_inicial__lte=bono.solicitud.fecha_autorizacion, fecha_final__gte=bono.solicitud.fecha_autorizacion).first()
        except AttributeError as e:
            catorcena = None
             
            
    total_monto = bonos.aggregate(total_monto=Sum('cantidad'))['total_monto']
    cantidad_bonos_aprobados = bonos.count()
    
    if request.method =='POST' and 'excel' in request.POST:
        try:
            return convert_excel_bonos_aprobados(bonos,catorcena,total_monto,cantidad_bonos_aprobados)
        except Exception as e:
                messages.error(request,'No existen bono aprobados para generar el reporte')
    
    p = Paginator(bonos, 50)
    page = request.GET.get('page')
    salidas_list = p.get_page(page)
    bonos = p.get_page(page)
        
    contexto = {
        'bonos':bonos,
        'bonosolicitado_filter':bonosolicitado_filter,
        'cantidad_bonos_aprobados':cantidad_bonos_aprobados,
        'catorcena':catorcena,
        'total_monto':total_monto,
        'salidas_list':salidas_list,
        'usuario':usuario,
        'ids':ids
    }
            
    return render(request,'esquema/bonos_varilleros/generar_reporte_bonos.html',contexto)

#para remover bonos agregados
@login_required(login_url='user-login')
@perfil_session_seleccionado
def removerBono(request,bono_id):
    #hacer el complete requerimiento a 0 - contar el numero de archivos cuando es 0
    if request.method == "POST":
        
        #obtener el usuario
        userdatos = request.session.get('usuario_datos')    
        usuario_id = userdatos.get('usuario_id')
        usuario = UserDatos.objects.get(pk = usuario_id)
        
        if usuario.tipo.id in (4,5):
            try:
                bono_solicitado = BonoSolicitado.objects.get(pk=bono_id)
                
                bandera = False #Se utiliza para saber si se realiza una division o suma acumulativa
                
                if bono_solicitado.bono.puesto.id == 7: # id puesto - todos los que participen en la actividad
                    #indica que se hara una division
                    bandera = True
                    solicitud_id = bono_solicitado.solicitud.id
                    bono_solicitado.delete()
                    #obtengo el total de la solicitud y ese total es divido entre en numero de empleados
                    total = Solicitud.objects.filter(pk = solicitud_id).values_list('total',flat=True) #recuerda acceder al valor como total[0]
                    participantes = BonoSolicitado.objects.filter(solicitud_id = solicitud_id).count()
                    if participantes > 0: #No se puede dividir entre 0
                        reparto = Decimal(total[0]/participantes)
                        BonoSolicitado.objects.filter(solicitud_id = solicitud_id).update(cantidad=reparto)
                        return JsonResponse({'bono_id': bono_id,'total':total[0], 'bandera':bandera,'reparto':reparto} ,status=200, safe=True)  
                    else:
                        return JsonResponse({'bono_id': bono_id,'total':0,'bandera':bandera,'reparto':0} ,status=200, safe=True)
                
                elif bono_solicitado.bono.esquema_subcategoria.id == 21: # Bono ó Subcategoria ID - TOMA DE REGISTROS ECO Y/O DINA - 9 soportes se paga, 30% ayudante, 70% tecnico
                    subcategoria = bono_solicitado.bono.esquema_subcategoria.id
                    #obtengo el id de la solicitud
                    solicitud_id = bono_solicitado.solicitud_id
                    BonoSolicitado.objects.filter(solicitud_id=solicitud_id).delete()
                    return JsonResponse({'bono_id': bono_id,'total':0.00, 'bandera':bandera,'subcategoria':subcategoria} ,status=200, safe=True)
                else:
                    #obtengo el id de la solicitud
                    solicitud_id = bono_solicitado.solicitud_id
                    #se elimina el registro del bonosolicitud
                    bono_solicitado.delete()
                    #se realiza la suma con los que queda en la BD
                    total = BonoSolicitado.objects.filter(solicitud_id = solicitud_id).aggregate(total=Sum('cantidad'))['total'] or 0
                    #se actualza la solicitud el total
                    Solicitud.objects.filter(pk = solicitud_id).update(total = total)  
                    return JsonResponse({'bono_id': bono_id,'total':total, 'bandera':bandera} ,status=200, safe=True)              
                
            except:
                return JsonResponse({'mensaje': 'No encontrado'}, status=404,safe=True)
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)

#para remover bonos verificar     
@login_required(login_url='user-login')
def removerBonoVerificar(request,bono_id):
    #hacer el complete requerimiento a 0 - contar el numero de archivos cuando es 0
    if request.method == "POST":
        
        usuario = get_object_or_404(UserDatos,user_id = request.user.id)
        
        if usuario.tipo.id in (4,5):
            try:
                bono = BonoSolicitado.objects.get(pk=bono_id)
                solicitud = Solicitud.objects.get(pk=bono.solicitud_id)
                if bono.puesto.id == 19: #ID puesto - todos los que participen en la actividad
                    bandera = 0 #se utiliza para ocultar la tabla del bono
                    participantes = BonoSolicitado.objects.filter(solicitud_id = bono.solicitud.id).count()#se cuenta el numero
                    bandera = participantes
                    participantes-=1 #se resta porque al final se le quita uno y se debe quedar la cuenta
                    solicitud = Solicitud.objects.get(pk=bono.solicitud.id)
                    if participantes == 0:
                        participantes = 1
                    reparto = Decimal(solicitud.total/participantes) #se hace la división
                    BonoSolicitado.objects.filter(solicitud_id = bono.solicitud.id).update(cantidad=reparto)
                    bono.delete()
                    return JsonResponse({'bono_id': bono_id,'total':solicitud.total, 'reparto':True, 'monto':reparto, 'participantes':participantes, 'bandera':bandera} ,status=200, safe=True)
                
                
                
                
                else:    
                    solicitud.total -= bono.cantidad
                    solicitud.save()
                    bono.delete()
                
                    return JsonResponse({
                        'bono_id': bono_id,
                        'total':solicitud.total, 
                        'reparto':False} ,status=200, safe=True)
            
            except:
                return JsonResponse({'mensaje': 'No encontrado'}, status=404,safe=True)
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)
        
        
#para remover archivos agregados
@login_required(login_url='user-login')
@perfil_session_seleccionado
def removerArchivo(request,archivo_id):
    if request.method == "POST":
        
        #obtener el usuario
        userdatos = request.session.get('usuario_datos')    
        usuario_id = userdatos.get('usuario_id')
        usuario = UserDatos.objects.get(pk = usuario_id)
        
        if usuario.tipo.id in (4,5):
            try:
                archivo = get_object_or_404(Requerimiento,pk=archivo_id)
                
                if os.path.isfile(archivo.url.path):
                    os.remove(archivo.url.path)
                    
                archivo.delete()
                            
                return JsonResponse({"archivo_id":archivo_id},status=200,safe=False)
            except:
                return JsonResponse({'mensaje':'objecto no encontrado'},status=404,safe=True)
            
        else:
            return JsonResponse({'mensaje': 'Prohibido'}, status=403,safe=True)        
        
#GENERACION DE REPORTES EN EXCEL
def convert_excel_bonos_aprobados(bonos,catorcena,total_monto,cantidad_bonos_aprobados):
    response= HttpResponse(content_type = "application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename = Reporte_bonos_varilleros_aprobados_' + str(date.today())+'.xlsx'
    wb = Workbook()
    ws = wb.create_sheet(title='Reporte')
    #Comenzar en la fila 1
    row_num = 1
    
    #Create heading style and adding to workbook | Crear el estilo del encabezado y agregarlo al Workbook
    head_style = NamedStyle(name = "head_style")
    head_style.font = Font(name = 'Arial', color = '00FFFFFF', bold = True, size = 11)
    head_style.fill = PatternFill("solid", fgColor = '00003366')
    wb.add_named_style(head_style)
    #Create body style and adding to workbook
    body_style = NamedStyle(name = "body_style")
    body_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(body_style)
    #Create messages style and adding to workbook
    messages_style = NamedStyle(name = "mensajes_style")
    messages_style.font = Font(name="Arial Narrow", size = 11)
    wb.add_named_style(messages_style)
    #Create date style and adding to workbook
    number_style = NamedStyle(name='number_style', number_format='#,##0')
    date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY HH:MM')
    date_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(date_style)
    money_style = NamedStyle(name='money_style', number_format='$ #,##0.00')
    money_style.font = Font(name ='Calibri', size = 10)
    wb.add_named_style(money_style)
    money_resumen_style = NamedStyle(name='money_resumen_style', number_format='$ #,##0.00')
    money_resumen_style.font = Font(name ='Calibri', size = 14, bold = True)
    wb.add_named_style(money_resumen_style)
    dato_style = NamedStyle(name='dato_style',number_format='DD/MM/YYYY')
    dato_style.font = Font(name="Arial Narrow", size = 11)
    
    
    
    #se crea el encabezado de la tabla en excel 
    columns = ['Folio','Fecha emisión','Fecha aprobación','# Trabajador','Nombre','No. de cuenta','No. de tarjeta','Banco','Distrito','Bono','Puesto','Cantidad']
    
    #se añade el ancho de cada columna
    for col_num in range(len(columns)):
        (ws.cell(row = row_num, column = col_num+1, value=columns[col_num])).style = head_style
        if col_num < 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 10
        if col_num == 4:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 15
            
    columna_max = len(columns)+2
    
    (ws.cell(column = columna_max, row = 1, value='{Reporte Creado Automáticamente por Savia RH. UH}')).style = messages_style
    (ws.cell(column = columna_max, row = 2, value='{Software desarrollado por Vordcab S.A. de C.V.}')).style = messages_style
    (ws.cell(column = columna_max, row = 3, value='')).style = messages_style
    (ws.cell(column = columna_max, row = 5, value=f'Catorcena: {catorcena.catorcena}: {catorcena.fecha_inicial} - {catorcena.fecha_final}')).style = dato_style
    (ws.cell(column = columna_max, row = 6, value=f'Bonos aprobados: {cantidad_bonos_aprobados}')).style = messages_style
    (ws.cell(column = columna_max, row = 7, value=f'Total $: {total_monto}')).style = messages_style
    ws.column_dimensions[get_column_letter(columna_max)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    ws.column_dimensions[get_column_letter(columna_max + 1)].width = 45
    
    rows = []
    
    #aqui se recorre el query de los bonos y se debe formatear los objectos a un tipo de dato
    for bono in bonos:

        row = (
            bono.solicitud.folio,
            bono.fecha.strftime('%Y-%m-%d %H:%M'),
            bono.solicitud.fecha_autorizacion.strftime('%Y-%m-%d %H:%M'),
            str(bono.trabajador.numero_de_trabajador),
            str(bono.trabajador),
            bono.trabajador.status.datosbancarios.no_de_cuenta,
            bono.trabajador.status.datosbancarios.numero_de_tarjeta,
            str(bono.trabajador.status.datosbancarios.banco),
            str(bono.bono.distrito),
            str(bono.solicitud.bono),
            str(bono.bono.puesto),
            bono.cantidad
        )
        rows.append(row)
        
        #aqui se empieza el recorrido para el llenado de datos de acuerdo a su tipo
        for row_num, row in enumerate(rows, start=2):
            for col_num, value in enumerate(row, start=1):
                if col_num == 1:
                    ws.cell(row=row_num, column=col_num, value=value).style = number_style
                elif col_num == 2 or col_num == 3:
                    ws.cell(row=row_num, column=col_num, value=value).style = date_style
                elif col_num <= 9:
                    ws.cell(row=row_num, column=col_num, value=value).style = body_style
                else:
                    ws.cell(row=row_num, column=col_num, value=value).style = money_style
    
    sheet = wb['Sheet']
    wb.remove(sheet)
    wb.save(response)
    
    return(response)

@login_required(login_url='user-login')
def tabuladorBonos(request):
    
    return render(request, 'esquema/crear_bonos/tabulador_bonos.html')

@login_required(login_url='user-login')
@perfil_session_seleccionado
def get_puestos(request):    
    try:
        #obtener el usuario
        userdatos = request.session.get('usuario_datos')    
        usuario_id = userdatos.get('usuario_id')
        usuario = UserDatos.objects.get(pk = usuario_id)
        
        bono_id = request.GET.get('bono_id')
        folio = request.GET.get('folio')
        data = []
        
        if bono_id:
            if int(bono_id) in (1,2):#IDS DEL MODELO ESQUEMA_SUBATEGORIA - evic extracion, evic introduccion
                bonos_solicitados = BonoSolicitado.objects.filter(solicitud__folio = folio).select_related('bono__puesto').values_list('bono__puesto_id',flat=True)
                #Trae todos los bonos por puesto por distrito, estado 1 = baja | 0 = activo - se excluye el puesto para este caso
                puestos_qs = Bono.objects.filter(esquema_subcategoria = bono_id,distrito_id = usuario.distrito.id,estado = 0).exclude(puesto_id__in = list(bonos_solicitados)).values('id','puesto__id','puesto__puesto','importe')
            else:
                #Trae todos los bonos por puesto por distrito, estado 1 = baja | 0 = activo
                puestos_qs = Bono.objects.filter(esquema_subcategoria = bono_id,distrito_id = usuario.distrito.id,estado = 0).values('id','puesto__id','puesto__puesto','importe')
            
            for puesto in puestos_qs:
                puesto['importe'] = str(puesto['importe'])
            data = list(puestos_qs)
            
        return JsonResponse(data, safe=False)
    
    except Exception as e:
        
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
